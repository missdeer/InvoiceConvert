#!/usr/bin/env python3
"""
Excel column conversion and deduplication tool.
Reads from one Excel file and writes to another with column mapping and aggregation.
"""

import sys
import os
import pandas as pd
import openpyxl
import re
import warnings
from pathlib import Path

# Suppress openpyxl style warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl.styles.stylesheet')

try:
    import pdfplumber
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False


def validate_input_file(input_file):
    """Validate input file before processing."""
    # Check if file exists
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"输入文件不存在: '{input_file}'")
    
    # Check if it's a file (not a directory)
    if not os.path.isfile(input_file):
        raise ValueError(f"输入路径不是文件: '{input_file}'")
    
    # Check if file is readable
    if not os.access(input_file, os.R_OK):
        raise PermissionError(f"输入文件不可读: '{input_file}'")
    
    # Check file extension
    if not input_file.lower().endswith(('.xlsx', '.xls')):
        raise ValueError(f"输入文件必须是Excel文件（.xlsx或.xls）: '{input_file}'")
    
    # Check file size (should not be empty)
    file_size = os.path.getsize(input_file)
    if file_size == 0:
        raise ValueError(f"输入文件为空: '{input_file}'")
    
    return True


def validate_output_file(output_file):
    """Validate output file path."""
    # Check if output directory exists and is writable
    output_dir = os.path.dirname(os.path.abspath(output_file))
    if output_dir and not os.path.exists(output_dir):
        try:
            os.makedirs(output_dir, exist_ok=True)
        except OSError as e:
            raise OSError(f"无法创建输出目录 '{output_dir}': {str(e)}")
    
    # Check if output directory is writable
    if output_dir and not os.access(output_dir, os.W_OK):
        raise PermissionError(f"输出目录不可写: '{output_dir}'")
    
    # Check file extension
    if not output_file.lower().endswith('.xlsx'):
        raise ValueError(f"输出文件必须是Excel文件（.xlsx）: '{output_file}'")
    
    return True


def extract_invoice_info_from_pdf(pdf_path):
    """
    Extract invoice information from text-based PDF file.
    
    Args:
        pdf_path: Path to the PDF file
        
    Returns:
        dict: Dictionary containing:
              - 'data': extracted invoice information
              - 'missing_fields': list of field names that failed to extract
              - 'error': error message if extraction completely failed
              Returns None if PDF cannot be opened or has no text
    """
    if not PDF_SUPPORT:
        return {'error': 'pdfplumber未安装', 'data': {}, 'missing_fields': []}
    
    invoice_info = {}
    missing_fields = []
    
    # Define required fields and their display names
    field_names = {
        'invoice_number': '发票号码',
        'invoice_amount': '开票金额',
        'tax_rate': '税率',
        'amount_excluding_tax': '不含税金额',
        'tax_amount': '税额',
        'invoice_date': '开票日期'
    }
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Extract text blocks with better context preservation
            # Use extract_text() which preserves layout, then split into meaningful blocks
            text_blocks = []
            full_text = ""
            
            for page in pdf.pages:
                # Extract text preserving layout
                page_text = page.extract_text(layout=True)
                if page_text:
                    page_text = page_text.strip()  # Trim page text
                    full_text += page_text + "\n"
                    # Split into lines and create blocks (group consecutive lines)
                    lines = page_text.split('\n')
                    current_block = []
                    for line in lines:
                        line = line.strip()
                        if line:
                            current_block.append(line)
                        else:
                            # Empty line indicates block boundary
                            if current_block:
                                block_text = ' '.join(current_block).strip()
                                if block_text:
                                    text_blocks.append(block_text)
                                current_block = []
                    # Add remaining block
                    if current_block:
                        block_text = ' '.join(current_block).strip()
                        if block_text:
                            text_blocks.append(block_text)
            
            # Trim full_text
            full_text = full_text.strip()
            
            # Create a comprehensive text for matching (use text blocks)
            # Join blocks with newlines to preserve context
            comprehensive_text = '\n'.join(text_blocks).strip()
            
            # Also keep the full text for fallback matching
            if not comprehensive_text:
                comprehensive_text = full_text
            
            if not comprehensive_text and not full_text:
                return {'error': 'PDF文件无文本内容（可能是扫描图像）', 'data': {}, 'missing_fields': list(field_names.values())}
            
            # Use comprehensive_text for matching (has better context)
            # Fallback to full_text if comprehensive_text is empty
            search_text = (comprehensive_text if comprehensive_text else full_text).strip()
            
            # Extract information from text_blocks first (more precise)
            # Extract invoice number (发票号码)
            # Format: "发票号码：25117000001187191134" (can be long numbers)
            # Look for blocks containing "发票号码"
            patterns_invoice_number = [
                r'发\s*票\s*号\s*码[：:]\s*(\d{8,25})',  # Match long invoice numbers (extend to 25 digits)
                r'发\s*票\s*号\s*码[：:\s]+(\d{8,25})',
                r'发\s*票\s*号\s*码[：:]\s*(\d+)',  # More flexible: match any digits after colon
                r'发\s*票\s*号\s*码[：:\s]+(\d+)',
                r'发\s*票\s*代\s*码[：:\s]+(\d{10,12})[^\d]*发\s*票\s*号\s*码[：:\s]+(\d{8,25})',
            ]
            extracted = False
            extracted_from_block = None
            # Search in text_blocks first
            for idx, block in enumerate(text_blocks, 1):
                for pattern_idx, pattern in enumerate(patterns_invoice_number, 1):
                    match = re.search(pattern, block)
                    if match:
                        # Get the last group (use groups()[-1] if multiple groups, otherwise group(1))
                        if len(match.groups()) > 1:
                            extracted_value = match.groups()[-1]
                        else:
                            extracted_value = match.group(1)
                        invoice_info['invoice_number'] = extracted_value
                        extracted = True
                        extracted_from_block = idx
                        break
                if extracted:
                    break
            if not extracted:
                for pattern_idx, pattern in enumerate(patterns_invoice_number, 1):
                    match = re.search(pattern, search_text, re.MULTILINE)
                    if match:
                        # Get the last group (use groups()[-1] if multiple groups, otherwise group(1))
                        if len(match.groups()) > 1:
                            invoice_info['invoice_number'] = match.groups()[-1]
                        else:
                            invoice_info['invoice_number'] = match.group(1)
                        extracted = True
                        break
            if not extracted:
                missing_fields.append(field_names['invoice_number'])
            
            # Extract invoice amount (开票金额/价税合计)
            # Format: "价税合计（大写）...（小写）¥66.20" or "价税合计（ 大写 ）...（ 小写 ）¥24.00"
            # Special format: "票价:￥58.00" - extract as invoice_amount, set tax_rate=9%, calculate amount_excluding_tax and tax_amount
            # Look for blocks containing "价税合计" and "小写", or "票价"
            patterns_amount = [
                r'（\s*小\s+写\s*）\s+[¥￥]?\s*([\d,]+\.?\d*)',  # Match "（ 小 写）" with spaces between 小 and 写
                r'（\s*小写\s*）\s+[¥￥]?\s*([\d,]+\.?\d*)',  # Primary pattern: match after "（小写）" with one or more spaces before ¥
                r'（\s*小\s*写\s*）\s*[¥￥]\s*([\d,]+\.?\d*)',  # Match "（ 小 写）" followed by optional spaces and ¥
                r'（\s*小写\s*）\s*[¥￥]\s*([\d,]+\.?\d*)',  # Match "（小写）" followed by optional spaces and ¥
                r'小\s+写\s*[）)）]\s+[¥￥]?\s*([\d,]+\.?\d*)',  # Match "小 写）" with spaces between 小 and 写
                r'小写\s*[）)）]\s+[¥￥]?\s*([\d,]+\.?\d*)',  # Match "小写）" or "小写 ）" with flexible spaces
                r'价税合计[（(（]\s*小\s+写\s*[）)）]\s*[：:\s]*[¥￥]?\s*([\d,]+\.?\d*)',  # Match with spaces between 小 and 写
                r'价税合计[（(（]\s*小写\s*[）)）]\s*[：:\s]*[¥￥]?\s*([\d,]+\.?\d*)',
                r'价税合计[：:\s]*[¥￥]?\s*([\d,]+\.?\d*)',
                r'（\s*小\s+写\s*）[^\d]*?([\d,]+\.?\d*)',  # Match "（ 小 写）" followed by any non-digit chars until number
                r'（\s*小写\s*）[^\d]*?([\d,]+\.?\d*)',  # Match "（小写）" followed by any non-digit chars until number
            ]
            # Pattern for "票价" format
            pattern_piao_jia = r'票价[：:]\s*[¥￥]?\s*([\d,]+\.?\d*)'
            extracted = False
            amount_from_block = None
            is_piao_jia_format = False  # Flag to indicate if we matched "票价" format
            # First, check for "票价" format (special handling)
            for idx, block in enumerate(text_blocks, 1):
                if '票价' in block:
                    match = re.search(pattern_piao_jia, block)
                    if match:
                        try:
                            amount_str = match.group(1).replace(',', '').replace('，', '').replace(' ', '')
                            amount_value = float(amount_str)
                            invoice_info['invoice_amount'] = amount_value
                            # Set tax rate to 9%
                            invoice_info['tax_rate'] = 9.0
                            # Calculate amount excluding tax: 不含税金额 = 开票金额 / (1 + 税率)
                            amount_excluding_tax = amount_value / 1.09
                            invoice_info['amount_excluding_tax'] = amount_excluding_tax
                            # Calculate tax amount: 税额 = 开票金额 - 不含税金额
                            tax_amount = amount_value - amount_excluding_tax
                            invoice_info['tax_amount'] = tax_amount
                            extracted = True
                            amount_from_block = idx
                            is_piao_jia_format = True
                            break
                        except ValueError as e:
                            continue
            # If not found "票价" format, search for normal patterns
            if not extracted:
                # Search in text_blocks first
                for idx, block in enumerate(text_blocks, 1):
                    has_jia_shui = '价税合计' in block
                    has_xiao_xie = '小写' in block
                    if has_jia_shui and has_xiao_xie:
                        for pattern_idx, pattern in enumerate(patterns_amount, 1):
                            match = re.search(pattern, block)
                            if match:
                                try:
                                    amount_str = match.group(1).replace(',', '').replace('，', '').replace(' ', '')
                                    amount_value = float(amount_str)
                                    invoice_info['invoice_amount'] = amount_value
                                    extracted = True
                                    amount_from_block = idx
                                    break
                                except ValueError as e:
                                    continue
                        if extracted:
                            break
            # Fallback to comprehensive text if not found in blocks
            if not extracted:
                for pattern in patterns_amount:
                    match = re.search(pattern, search_text, re.MULTILINE | re.DOTALL)
                    if match:
                        try:
                            amount_str = match.group(1).replace(',', '').replace('，', '').replace(' ', '')
                            invoice_info['invoice_amount'] = float(amount_str)
                            extracted = True
                            break
                        except ValueError:
                            continue
            if not extracted:
                missing_fields.append(field_names['invoice_amount'])
            
            # Extract tax rate (税率)
            # Format: Can be "3%" in a text block (like block 2), or "税率：3%"
            patterns_tax_rate = [
                r'\b(\d+\.?\d*)%',  # Match standalone percentage like "3%"
            ]
            extracted = False
            rate_from_block = None
            # Skip if "票价" format was already matched (tax_rate is already set to 9%)
            if is_piao_jia_format:
                extracted = True
            # Search in text_blocks first (look for blocks with percentage)
            if not extracted:
                for idx, block in enumerate(text_blocks, 1):
                    if '%' in block:
                        for pattern_idx, pattern in enumerate(patterns_tax_rate, 1):
                            matches = re.findall(pattern, block)
                            if matches:
                                # Filter reasonable tax rates (typically 0-100%)
                                for match in matches:
                                    try:
                                        rate_value = float(match)
                                        # Tax rates are typically between 0 and 100
                                        # Common rates: 0%, 3%, 6%, 9%, 13%, 16%, etc.
                                        if 0 <= rate_value <= 100:
                                            invoice_info['tax_rate'] = rate_value
                                            extracted = True
                                            rate_from_block = idx
                                            break
                                    except ValueError as e:
                                        continue
                            if extracted:
                                break
                        if extracted:
                            break
            # Fallback to comprehensive text if not found in blocks
            if not extracted:
                for pattern in patterns_tax_rate:
                    matches = re.findall(pattern, search_text, re.MULTILINE | re.DOTALL)
                    if matches:
                        for match in matches:
                            try:
                                rate_value = float(match)
                                if 0 <= rate_value <= 100:
                                    invoice_info['tax_rate'] = rate_value
                                    extracted = True
                                    break
                            except ValueError:
                                continue
                    if extracted:
                        break
            # If still not extracted, set default value to NaN
            if not extracted:
                import math
                invoice_info['tax_rate'] = float('nan')
            
            # Extract amount excluding tax (不含税金额)
            # Format: "合     计                                      ¥64.28                  ¥1.92"
            # Look for blocks containing "合     计" (with spaces) and extract the first ¥ amount
            # Use '合\s+计' to avoid matching "价税合计"
            # Note: If "票价" format was matched, amount_excluding_tax is already set, skip extraction
            patterns_excl_tax = [
                r'合\s+计[^\¥￥]*?[¥￥]\s*([\d,]+\.?\d*)',  # Match first ¥ amount after "合     计"
                r'不含税金额[：:\s]*[¥￥]?\s*([\d,]+\.?\d*)',
                r'合计金额[：:\s]*[¥￥]?\s*([\d,]+\.?\d*)',
            ]
            extracted = False
            excl_tax_from_block = None
            # Skip if "票价" format was already matched (amount_excluding_tax is already set)
            if is_piao_jia_format:
                extracted = True
            # Search in text_blocks first
            if not extracted:
                for idx, block in enumerate(text_blocks, 1):
                    # Check for "合" followed by spaces and then "计" (not "价税合计")
                    has_he_ji = re.search(r'合\s+计', block) is not None
                    has_yuan = '¥' in block
                    if has_he_ji and has_yuan:
                        # Find amounts after "合\s+计" but before "价税合计" (if present)
                        # Extract the portion from "合\s+计" to "价税合计" (or end of block)
                        he_ji_match = re.search(r'合\s+计', block)
                        if he_ji_match:
                            start_pos = he_ji_match.end()
                            # Find "价税合计" position if it exists
                            jia_shui_match = re.search(r'价税合计', block)
                            if jia_shui_match and jia_shui_match.start() > start_pos:
                                # Only extract amounts between "合\s+计" and "价税合计"
                                relevant_text = block[start_pos:jia_shui_match.start()]
                            else:
                                # Extract amounts after "合\s+计" to end of block
                                relevant_text = block[start_pos:]
                            # Find all amounts in the relevant portion
                            amounts = re.findall(r'[¥￥][\s]*([\d,]+\.?\d*)', relevant_text)
                            
                            # Also try a more comprehensive pattern
                            if len(amounts) < 1:
                                amounts = re.findall(r'[¥￥][^\d]*?([\d,]+\.?\d*)', relevant_text)
                        else:
                            # Fallback: find all amounts in the block
                            amounts = re.findall(r'[¥￥][\s]*([\d,]+\.?\d*)', block)
                            if len(amounts) < 1:
                                amounts = re.findall(r'[¥￥][^\d]*?([\d,]+\.?\d*)', block)
                        
                        if len(amounts) >= 1:
                            # First amount is the amount excluding tax
                            try:
                                amount_str = amounts[0].replace(',', '').replace('，', '').replace(' ', '').strip()
                                amount_value = float(amount_str)
                                invoice_info['amount_excluding_tax'] = amount_value
                                extracted = True
                                excl_tax_from_block = idx
                                break
                            except (ValueError, IndexError) as e:
                                pass
                        # Try pattern matching as fallback
                        if not extracted:
                            for pattern_idx, pattern in enumerate(patterns_excl_tax, 1):
                                match = re.search(pattern, block)
                                if match:
                                    try:
                                        amount_str = match.group(1).replace(',', '').replace('，', '').replace(' ', '').strip()
                                        amount_value = float(amount_str)
                                        invoice_info['amount_excluding_tax'] = amount_value
                                        extracted = True
                                        excl_tax_from_block = idx
                                        break
                                    except ValueError as e:
                                        continue
                        if extracted:
                            break
            # Fallback to comprehensive text if not found in blocks
            if not extracted:
                for pattern in patterns_excl_tax:
                    match = re.search(pattern, search_text, re.MULTILINE | re.DOTALL)
                    if match:
                        try:
                            amount_str = match.group(1).replace(',', '').replace('，', '').replace(' ', '').strip()
                            invoice_info['amount_excluding_tax'] = float(amount_str)
                            extracted = True
                            break
                        except ValueError:
                            continue
            if not extracted:
                missing_fields.append(field_names['amount_excluding_tax'])
            
            # Extract tax amount (税额)
            # Format: "合     计                                      ¥64.28                  ¥1.92"
            # Look for blocks containing "合     计" (with spaces) and extract the second ¥ amount
            # Use '合\s+计' to avoid matching "价税合计"
            # Note: If "票价" format was matched, tax_amount is already set, skip extraction
            patterns_tax_amount = [
                r'合\s+计[^\¥￥]*?[¥￥]\s*[\d,]+\.?\d*[^\¥￥]*?[¥￥]\s*([\d,]+\.?\d*)',  # Match second ¥ amount after "合     计"
            ]
            extracted = False
            tax_amount_from_block = None
            # Skip if "票价" format was already matched (tax_amount is already set)
            if is_piao_jia_format:
                extracted = True
            # Search in text_blocks first
            if not extracted:
                for idx, block in enumerate(text_blocks, 1):
                    # Check for "合" followed by spaces and then "计" (not "价税合计")
                    has_he_ji = re.search(r'合\s+计', block) is not None
                    has_yuan = '¥' in block
                    if has_he_ji and has_yuan:
                        # Find amounts after "合\s+计" but before "价税合计" (if present)
                        # 税额必须是"合\s+计"所在行的第2个¥之后的数字
                        # 如果有"价税合计"在同一行，则要看第2个¥是在"价税合计"前还是后
                        # 如果是前，则提取；如果是后，则不是税额，设为0.0
                        he_ji_match = re.search(r'合\s+计', block)
                        if he_ji_match:
                            start_pos = he_ji_match.end()
                            # Find "价税合计" position if it exists
                            jia_shui_match = re.search(r'价税合计', block)
                            if jia_shui_match and jia_shui_match.start() > start_pos:
                                # Only extract amounts between "合\s+计" and "价税合计"
                                relevant_text = block[start_pos:jia_shui_match.start()]
                            else:
                                # Extract amounts after "合\s+计" to end of block
                                relevant_text = block[start_pos:]
                            
                            # Find all ¥ positions in the relevant portion
                            # Use finditer to get positions
                            yuan_matches = list(re.finditer(r'[¥￥]', relevant_text))
                            
                            if len(yuan_matches) >= 2:
                                # Check if the second ¥ is before "价税合计"
                                second_yuan_pos = yuan_matches[1].start()
                                if jia_shui_match and jia_shui_match.start() > start_pos:
                                    # Calculate absolute position in block
                                    second_yuan_abs_pos = start_pos + second_yuan_pos
                                    jia_shui_abs_pos = jia_shui_match.start()
                                    if second_yuan_abs_pos < jia_shui_abs_pos:
                                        # Second ¥ is before "价税合计", extract it
                                        # Extract amount after the second ¥
                                        second_yuan_text = relevant_text[second_yuan_pos:]
                                        amount_match = re.search(r'[¥￥][\s]*([\d,]+\.?\d*)', second_yuan_text)
                                        if amount_match:
                                            try:
                                                amount_str = amount_match.group(1).replace(',', '').replace('，', '').replace(' ', '').strip()
                                                amount_value = float(amount_str)
                                                invoice_info['tax_amount'] = amount_value
                                                extracted = True
                                                tax_amount_from_block = idx
                                                break
                                            except (ValueError, IndexError) as e:
                                                pass
                                else:
                                    # No "价税合计" in this block, extract second amount
                                    second_yuan_text = relevant_text[second_yuan_pos:]
                                    amount_match = re.search(r'[¥￥][\s]*([\d,]+\.?\d*)', second_yuan_text)
                                    if amount_match:
                                        try:
                                            amount_str = amount_match.group(1).replace(',', '').replace('，', '').replace(' ', '').strip()
                                            amount_value = float(amount_str)
                                            invoice_info['tax_amount'] = amount_value
                                            extracted = True
                                            tax_amount_from_block = idx
                                            break
                                        except (ValueError, IndexError) as e:
                                            pass
                        if extracted:
                            break
            # If still not extracted, set default value to 0.0
            # 税额必须从"合\s+计"后的第2个¥提取，如果没有找到，则设为0.0
            if not extracted:
                invoice_info['tax_amount'] = 0.0
            
            # Extract invoice date (开票日期)
            # Format: "开票日期：2025年09月23日"
            # Look for blocks containing "开票日期"
            patterns_date = [
                r'开\s*票\s*日\s*期[：:]\s*(\d{4})年(\d{1,2})月(\d{1,2})日',  # Primary pattern: YYYY年MM月DD日
                r'开\s*票\s*日\s*期[：:\s]*(\d{4})[-年](\d{1,2})[-月](\d{1,2})[日]?',
                r'开\s*票\s*日\s*期[：:\s]*(\d{4})[/](\d{1,2})[/](\d{1,2})',
            ]
            extracted = False
            date_from_block = None
            # Search in text_blocks first
            for idx, block in enumerate(text_blocks, 1):
                for pattern_idx, pattern in enumerate(patterns_date, 1):
                    match = re.search(pattern, block)
                    if match:
                        year, month, day = match.groups()
                        date_str = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                        invoice_info['invoice_date'] = date_str
                        extracted = True
                        date_from_block = idx
                        break
                if extracted:
                    break
            if not extracted:
                for pattern_idx, pattern in enumerate(patterns_date, 1):
                    match = re.search(pattern, search_text, re.MULTILINE | re.DOTALL)
                    if match:
                        year, month, day = match.groups()
                        date_str = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                        invoice_info['invoice_date'] = date_str
                        extracted = True
                        break
            if not extracted:
                missing_fields.append(field_names['invoice_date'])
                
    except Exception as e:
        return {'error': f'读取PDF文件失败: {str(e)}', 'data': {}, 'missing_fields': list(field_names.values())}
    
    return {
        'data': invoice_info,
        'missing_fields': missing_fields,
        'error': None
    }


def find_pdf_by_invoice_number(directory, invoice_number, recursive=False):
    """
    Find PDF file by invoice number.
    
    Supports multiple naming patterns:
    - {invoice_number}.pdf
    - {invoice_number}_*.pdf
    - *_{invoice_number}.pdf
    - *{invoice_number}*.pdf
    
    Args:
        directory: Directory to search in
        invoice_number: Invoice number to search for
        recursive: If True, recursively search all subdirectories
        
    Returns:
        str: Path to PDF file, or None if not found
    """
    if not directory or not os.path.exists(directory):
        return None
    
    pdf_dir = Path(directory)
    invoice_str = str(invoice_number).strip()
    
    # Determine search pattern based on recursive flag
    if recursive:
        # Recursive search: use **/ pattern
        search_patterns = [
            f"**/{invoice_str}.pdf",
            f"**/{invoice_str}_*.pdf",
            f"**/*_{invoice_str}.pdf",
            f"**/*{invoice_str}*.pdf",
        ]
    else:
        # Non-recursive: try exact match first
        exact_match = pdf_dir / f"{invoice_str}.pdf"
        if exact_match.exists():
            return str(exact_match)
        
        # Then try patterns with wildcards (non-recursive)
        search_patterns = [
            f"{invoice_str}_*.pdf",
            f"*_{invoice_str}.pdf",
            f"*{invoice_str}*.pdf",
        ]
    
    # Search using patterns
    for pattern in search_patterns:
        matches = list(pdf_dir.glob(pattern))
        if matches:
            # Return the first match
            return str(matches[0])
    
    return None


def find_pdf_directory(input_file):
    """
    Find PDF directory automatically.
    First check for 'pdfs' subdirectory, then check the same directory as input file.
    
    Args:
        input_file: Path to input Excel file
        
    Returns:
        str: Path to PDF directory, or None if not found
    """
    input_dir = os.path.dirname(os.path.abspath(input_file))
    if not input_dir:
        input_dir = os.getcwd()
    
    # First try: check for 'pdfs' subdirectory
    pdfs_subdir = os.path.join(input_dir, 'pdfs')
    if os.path.exists(pdfs_subdir) and os.path.isdir(pdfs_subdir):
        return pdfs_subdir
    
    # Second try: check the same directory as input file
    if os.path.exists(input_dir) and os.path.isdir(input_dir):
        # Check if there are any PDF files in this directory
        pdf_files = list(Path(input_dir).glob('*.pdf'))
        if pdf_files:
            return input_dir
    
    return None


def verify_excel_with_pdf(result_df, pdf_directory=None, recursive=False):
    """
    Verify Excel data against PDF files.
    
    Args:
        result_df: DataFrame with columns:
                   - Column 8 (I): 发票号码 (invoice_number)
                   - Column 3 (D): 开票金额 (invoice_amount)
                   - Column 4 (E): 税率 (tax_rate)
                   - Column 5 (F): 不含税金额 (amount_excluding_tax)
                   - Column 6 (G): 税额 (tax_amount)
        pdf_directory: Directory containing PDF files. If None, will not verify.
        recursive: If True, recursively search all subdirectories for PDF files.
        
    Returns:
        list: List of verification results
    """
    if not PDF_SUPPORT:
        print("提示: pdfplumber未安装，跳过PDF验证。运行 'pip install pdfplumber' 以启用PDF验证功能。")
        return []
    
    if result_df.empty:
        return []
    
    if pdf_directory is None:
        return []
    
    if not os.path.exists(pdf_directory):
        print(f"警告: PDF目录不存在: '{pdf_directory}'，跳过PDF验证")
        return []
    
    verification_results = []
    invoice_number_col = 8  # Column I (发票号码)
    invoice_amount_col = 3  # Column D (开票金额)
    tax_rate_col = 4        # Column E (税率)
    excl_tax_col = 5        # Column F (不含税金额)
    tax_amount_col = 6      # Column G (税额)
    
    search_mode = "递归搜索" if recursive else "仅当前目录"
    print(f"\n开始PDF验证 (PDF目录: '{pdf_directory}', 搜索模式: {search_mode})...")
    print(f"共 {len(result_df)} 条记录需要验证")
    
    for idx, row in result_df.iterrows():
        invoice_number = row[invoice_number_col]
        
        # Skip if invoice number is missing
        if pd.isna(invoice_number) or invoice_number == '':
            verification_results.append({
                'row': idx + 2,  # +2 because of header row
                'invoice_number': None,
                'status': 'SKIPPED',
                'message': '发票号码为空'
            })
            continue
        
        invoice_number_str = str(invoice_number).strip()
        
        # Find PDF file
        pdf_path = find_pdf_by_invoice_number(pdf_directory, invoice_number_str, recursive=recursive)
        
        if not pdf_path:
            verification_results.append({
                'row': idx + 2,
                'invoice_number': invoice_number_str,
                'status': 'PDF_NOT_FOUND',
                'message': f'未找到对应的PDF文件'
            })
            continue
        
        # Extract information from PDF
        pdf_extraction_result = extract_invoice_info_from_pdf(pdf_path)
        
        # Check if extraction completely failed
        if pdf_extraction_result is None:
            verification_results.append({
                'row': idx + 2,
                'invoice_number': invoice_number_str,
                'status': 'PDF_EXTRACTION_FAILED',
                'message': '无法打开PDF文件或PDF文件损坏',
                'missing_fields': []
            })
            continue
        
        # Check for extraction errors
        if pdf_extraction_result.get('error'):
            error_msg = pdf_extraction_result.get('error', '未知错误')
            missing_fields = pdf_extraction_result.get('missing_fields', [])
            verification_results.append({
                'row': idx + 2,
                'invoice_number': invoice_number_str,
                'status': 'PDF_EXTRACTION_FAILED',
                'message': error_msg,
                'missing_fields': missing_fields
            })
            continue
        
        # Get extracted data
        pdf_info = pdf_extraction_result.get('data', {})
        missing_fields = pdf_extraction_result.get('missing_fields', [])
        
        # If all critical fields are missing, mark as extraction failed
        critical_fields = ['开票金额', '税率', '不含税金额', '税额']
        missing_critical = [f for f in missing_fields if f in critical_fields]
        
        if len(missing_critical) == len(critical_fields):
            # All critical fields are missing, cannot verify
            verification_results.append({
                'row': idx + 2,
                'invoice_number': invoice_number_str,
                'status': 'PDF_EXTRACTION_FAILED',
                'message': f'无法提取关键字段: {", ".join(missing_critical)}',
                'missing_fields': missing_fields
            })
            continue
        
        # Compare data
        discrepancies = []
        tolerance = 0.01  # Allow 0.01 difference for floating point comparison
        
        # Verify invoice number
        pdf_invoice_number = pdf_info.get('invoice_number')
        if pdf_invoice_number and pdf_invoice_number != invoice_number_str:
            discrepancies.append(f"发票号码: Excel={invoice_number_str}, PDF={pdf_invoice_number}")
        
        # Verify invoice amount
        excel_amount = pd.to_numeric(row[invoice_amount_col], errors='coerce')
        pdf_amount = pdf_info.get('invoice_amount')
        if pdf_amount is not None and not pd.isna(excel_amount):
            if abs(float(excel_amount) - float(pdf_amount)) > tolerance:
                discrepancies.append(f"开票金额: Excel={excel_amount:.2f}, PDF={pdf_amount:.2f}")
        elif '开票金额' in missing_fields:
            discrepancies.append(f"开票金额: Excel={excel_amount:.2f}, PDF=无法提取")
        
        # Verify tax rate
        excel_rate = pd.to_numeric(row[tax_rate_col], errors='coerce')
        pdf_rate = pdf_info.get('tax_rate')
        if pdf_rate is not None:
            import math
            if math.isnan(pdf_rate):
                # PDF税率为NaN，不进行验证
                pass
            elif not pd.isna(excel_rate):
                if abs(float(excel_rate) - float(pdf_rate)) > tolerance:
                    discrepancies.append(f"税率: Excel={excel_rate:.2f}%, PDF={pdf_rate:.2f}%")
        elif '税率' in missing_fields:
            discrepancies.append(f"税率: Excel={excel_rate:.2f}%, PDF=无法提取")
        
        # Verify amount excluding tax
        excel_excl = pd.to_numeric(row[excl_tax_col], errors='coerce')
        pdf_excl = pdf_info.get('amount_excluding_tax')
        if pdf_excl is not None and not pd.isna(excel_excl):
            if abs(float(excel_excl) - float(pdf_excl)) > tolerance:
                discrepancies.append(f"不含税金额: Excel={excel_excl:.2f}, PDF={pdf_excl:.2f}")
        elif '不含税金额' in missing_fields:
            discrepancies.append(f"不含税金额: Excel={excel_excl:.2f}, PDF=无法提取")
        
        # Verify tax amount
        excel_tax = pd.to_numeric(row[tax_amount_col], errors='coerce')
        pdf_tax = pdf_info.get('tax_amount')
        if pdf_tax is not None and not pd.isna(excel_tax):
            if abs(float(excel_tax) - float(pdf_tax)) > tolerance:
                discrepancies.append(f"税额: Excel={excel_tax:.2f}, PDF={pdf_tax:.2f}")
        elif '税额' in missing_fields:
            discrepancies.append(f"税额: Excel={excel_tax:.2f}, PDF=无法提取")
        
        if discrepancies:
            verification_results.append({
                'row': idx + 2,
                'invoice_number': invoice_number_str,
                'status': 'MISMATCH',
                'discrepancies': discrepancies,
                'pdf_path': pdf_path,
                'missing_fields': missing_fields
            })
        else:
            # Check if there are missing fields (but verification still passed for available fields)
            if missing_fields:
                verification_results.append({
                    'row': idx + 2,
                    'invoice_number': invoice_number_str,
                    'status': 'MATCH',
                    'message': '验证通过（部分字段无法提取，但已提取字段匹配）',
                    'missing_fields': missing_fields
                })
            else:
                verification_results.append({
                    'row': idx + 2,
                    'invoice_number': invoice_number_str,
                    'status': 'MATCH',
                    'message': '验证通过'
                })
    
    # Print summary
    total = len(verification_results)
    matched = sum(1 for r in verification_results if r['status'] == 'MATCH')
    mismatched = sum(1 for r in verification_results if r['status'] == 'MISMATCH')
    not_found = sum(1 for r in verification_results if r['status'] == 'PDF_NOT_FOUND')
    failed = sum(1 for r in verification_results if r['status'] == 'PDF_EXTRACTION_FAILED')
    skipped = sum(1 for r in verification_results if r['status'] == 'SKIPPED')
    
    print(f"\nPDF验证完成:")
    print(f"  总计: {total}")
    print(f"  ✓ 匹配: {matched}")
    print(f"  ✗ 不匹配: {mismatched}")
    print(f"  ? PDF未找到: {not_found}")
    print(f"  ! 提取失败: {failed}")
    if skipped > 0:
        print(f"  - 跳过: {skipped}")
    
    # Print PDF not found errors
    if not_found > 0:
        print(f"\n错误: 找不到PDF文件的记录:")
        for result in verification_results:
            if result['status'] == 'PDF_NOT_FOUND':
                print(f"  行 {result['row']}, 发票号码: {result['invoice_number']} - {result['message']}")
    
    # Print extraction failed errors
    if failed > 0:
        print(f"\n错误: PDF信息提取失败的记录:")
        for result in verification_results:
            if result['status'] == 'PDF_EXTRACTION_FAILED':
                print(f"  行 {result['row']}, 发票号码: {result['invoice_number']}")
                print(f"    错误: {result['message']}")
                missing_fields = result.get('missing_fields', [])
                if missing_fields:
                    print(f"    无法提取的字段: {', '.join(missing_fields)}")
    
    # Print mismatches
    if mismatched > 0:
        print(f"\n不匹配的记录:")
        for result in verification_results:
            if result['status'] == 'MISMATCH':
                print(f"  行 {result['row']}, 发票号码: {result['invoice_number']}")
                for disc in result['discrepancies']:
                    print(f"    - {disc}")
                missing_fields = result.get('missing_fields', [])
                if missing_fields:
                    print(f"    注意: 以下字段无法从PDF提取: {', '.join(missing_fields)}")
    
    return verification_results


def process_excel(input_file, output_file, pdf_directory=None, pdf_recursive=False):
    """
    Process Excel file with column mapping and deduplication.
    
    Column mappings (from input to output):
    - D column (index 3) -> I column (index 8)
    - I column (index 8) -> B column (index 1)
    - T column (index 19) -> D column (index 3)
    - R column (index 17) -> E column (index 4)
    - Q column (index 16) -> F column (index 5)
    - S column (index 18) -> G column (index 6)
    - V column (index 21) -> H column (index 7)
    
    Special handling: If D column has duplicate values, aggregate Q, S, T columns
    by summing them and merge into a single row.
    """
    # Validate input file
    try:
        validate_input_file(input_file)
        print(f"✓ 输入文件验证通过: '{input_file}'")
    except (FileNotFoundError, ValueError, PermissionError) as e:
        print(f"错误: {str(e)}")
        sys.exit(1)
    
    # Validate output file
    try:
        validate_output_file(output_file)
        print(f"✓ 输出文件验证通过: '{output_file}'")
    except (OSError, ValueError, PermissionError) as e:
        print(f"错误: {str(e)}")
        sys.exit(1)
    
    # Read the input Excel file starting from row 2 (skip first row)
    # Only read the "信息汇总表" worksheet, ignore other sheets
    try:
        print(f"正在读取输入文件: '{input_file}' (工作表: '信息汇总表', 从第2行开始)...")
        # Skip the first row (index 0), read from row 2 onwards
        # Only read the "信息汇总表" worksheet
        df = pd.read_excel(input_file, engine='openpyxl', sheet_name='信息汇总表', header=None, skiprows=1)
        print(f"✓ 成功读取工作表 '信息汇总表'")
    except FileNotFoundError:
        print(f"错误: 输入文件未找到: '{input_file}'")
        sys.exit(1)
    except PermissionError:
        print(f"错误: 读取文件时权限被拒绝: '{input_file}'")
        sys.exit(1)
    except ValueError as e:
        error_msg = str(e)
        if 'Worksheet named' in error_msg or '工作表' in error_msg or 'sheet' in error_msg.lower():
            print(f"错误: 在文件 '{input_file}' 中未找到工作表 '信息汇总表'")
            print("请确保Excel文件包含名为 '信息汇总表' 的工作表。")
        else:
            print(f"错误: 读取Excel文件失败 '{input_file}': {error_msg}")
        sys.exit(1)
    except Exception as e:
        print(f"错误: 读取Excel文件失败 '{input_file}': {str(e)}")
        print("请确保文件是有效的Excel文件（.xlsx格式）且未损坏。")
        sys.exit(1)
    
    # Validate file content
    if df.empty:
        print("错误: 输入文件在跳过第一行后不包含数据行。")
        sys.exit(1)
    
    # Use df directly as data_df since we already skipped the first row
    data_df = df
    
    if data_df.empty:
        print("警告: 未找到数据行。输出文件将只包含表头行。")
    
    # Check if we have minimum required columns
    min_required_cols = 4  # At least need column D (index 3)
    if len(data_df.columns) < min_required_cols and not data_df.empty:
        print(f"警告: 输入文件只有 {len(data_df.columns)} 列。可能缺少某些列。")
    
    # Ensure we have enough columns (at least up to column V, which is index 21)
    # Excel columns: A=0, B=1, C=2, D=3, E=4, F=5, G=6, H=7, I=8, ..., Q=16, R=17, S=18, T=19, V=21
    max_col_index = 21
    if len(data_df.columns) <= max_col_index:
        # Add empty columns if needed
        for i in range(len(data_df.columns), max_col_index + 1):
            data_df[i] = None
    
    # Use data_df for processing instead of df
    df = data_df
    
    # Define column indices
    # Input columns
    D_col = 3
    I_col = 8
    T_col = 19
    R_col = 17
    Q_col = 16
    S_col = 18
    V_col = 21
    
    # Calculate original sums for Q, S, T columns before processing
    if not df.empty:
        original_q_sum = pd.to_numeric(df[Q_col], errors='coerce').sum()
        original_s_sum = pd.to_numeric(df[S_col], errors='coerce').sum()
        original_t_sum = pd.to_numeric(df[T_col], errors='coerce').sum()
        print(f"原始累加值 - Q列: {original_q_sum:.2f}, S列: {original_s_sum:.2f}, T列: {original_t_sum:.2f}")
    else:
        original_q_sum = 0
        original_s_sum = 0
        original_t_sum = 0
    
    if df.empty:
        print("没有数据行需要处理。")
        aggregated_df = pd.DataFrame()
    else:
        print(f"找到 {len(df)} 行数据需要处理")
        
        # Group by D column to handle duplicates
        # Convert D column values to string for grouping (handles NaN values)
        df['_group_key'] = df[D_col].fillna('').astype(str)
        
        # Aggregate rows with same D column value
        def aggregate_group(group):
            if len(group) == 1:
                return group.iloc[0]
            else:
                # Multiple rows with same D value - aggregate Q, S, T columns
                result = group.iloc[0].copy()
                
                # Sum Q, S, T columns (convert to numeric first, handling non-numeric values)
                q_sum = pd.to_numeric(group[Q_col], errors='coerce').sum()
                s_sum = pd.to_numeric(group[S_col], errors='coerce').sum()
                t_sum = pd.to_numeric(group[T_col], errors='coerce').sum()
                
                # Update Q, S, T with summed values
                result[Q_col] = q_sum if not pd.isna(q_sum) else None
                result[S_col] = s_sum if not pd.isna(s_sum) else None
                result[T_col] = t_sum if not pd.isna(t_sum) else None
                
                return result
        
        # Group by D column and aggregate
        try:
            aggregated_df = df.groupby('_group_key', as_index=False).apply(aggregate_group, include_groups=False).reset_index(drop=True)
            aggregated_df = aggregated_df.drop(columns=['_group_key'])
            print(f"✓ 成功处理并聚合数据")
        except Exception as e:
            print(f"错误: 处理数据失败: {str(e)}")
            sys.exit(1)
    
    # Create output dataframe with column mapping
    num_rows = len(aggregated_df)
    result_df = pd.DataFrame(index=range(num_rows))
    
    # Map columns according to original requirements
    # Original mapping: D->I, I->B, T->D, R->E, Q->F, S->G, V->H
    # Output header: A(序号), B(日期), C(所属类型), D(开票金额), E(税率), F(不含税金额), G(税额), H(发票种类), I(发票号码), J(报销人)
    # Following original mapping exactly:
    # I -> B: 日期 (column index 8 -> 1)
    # Convert datetime format (yyyy-MM-dd hh:mm:ss) to date format (yyyy-MM-dd)
    if I_col in aggregated_df.columns:
        try:
            # Convert to datetime and then format as date string
            date_values = pd.to_datetime(aggregated_df[I_col], errors='coerce')
            # Format as yyyy-MM-dd, handling NaT (Not a Time) values
            formatted_dates = date_values.dt.strftime('%Y-%m-%d')
            # Replace NaT with None
            formatted_dates = formatted_dates.where(pd.notna(formatted_dates), None)
            result_df[1] = formatted_dates.tolist()
        except Exception as e:
            print(f"警告: 转换I列日期格式失败: {str(e)}")
            print("使用原始值...")
            result_df[1] = aggregated_df[I_col].values
    else:
        result_df[1] = [None] * num_rows
    
    # T -> D: 开票金额 (column index 19 -> 3)
    if T_col in aggregated_df.columns:
        result_df[3] = aggregated_df[T_col].values
    else:
        result_df[3] = [None] * num_rows
    
    # R -> E: 税率 (column index 17 -> 4)
    if R_col in aggregated_df.columns:
        result_df[4] = aggregated_df[R_col].values
    else:
        result_df[4] = [None] * num_rows
    
    # Q -> F: 不含税金额 (column index 16 -> 5)
    if Q_col in aggregated_df.columns:
        result_df[5] = aggregated_df[Q_col].values
    else:
        result_df[5] = [None] * num_rows
    
    # S -> G: 税额 (column index 18 -> 6)
    if S_col in aggregated_df.columns:
        result_df[6] = aggregated_df[S_col].values
    else:
        result_df[6] = [None] * num_rows
    
    # V -> H: 发票种类 (column index 21 -> 7)
    if V_col in aggregated_df.columns:
        result_df[7] = aggregated_df[V_col].values
    else:
        result_df[7] = [None] * num_rows
    
    # D -> I: 发票号码 (column index 3 -> 8)
    if D_col in aggregated_df.columns:
        result_df[8] = aggregated_df[D_col].values
    else:
        result_df[8] = [None] * num_rows
    
    # Ensure all columns from A to J exist (fill with None if missing)
    for col_idx in range(10):
        if col_idx not in result_df.columns:
            result_df[col_idx] = [None] * num_rows
    
    # Add sequence number column (A column, index 0)
    # Sequence numbers start from 1, will be in row 2 (after header row)
    result_df[0] = [i + 1 for i in range(num_rows)]
    
    # Reorder columns to match Excel column order A-J (0-9)
    column_order = list(range(10))
    result_df = result_df[column_order]
    
    # Create output dataframe with specified header row
    # Header: 序号、日期、所属类型、开票金额、税率、不含税金额、税额、发票种类、发票号码、报销人
    # Columns: A(0), B(1), C(2), D(3), E(4), F(5), G(6), H(7), I(8), J(9)
    header_values = ['序号', '日期', '所属类型', '开票金额', '税率', '不含税金额', '税额', '发票种类', '发票号码', '报销人']
    
    # Ensure result_df has all columns from A to J (0 to 9)
    for col_idx in range(10):
        if col_idx not in result_df.columns:
            result_df[col_idx] = [None] * num_rows
    
    # Reorder result_df columns to A-J (0-9)
    result_df = result_df[list(range(10))]
    
    # Create header dataframe
    output_header = pd.DataFrame([header_values], columns=list(range(10)))
    
    # Combine header and data rows
    final_df = pd.concat([output_header, result_df], ignore_index=True)
    
    # Handle output file: create new or overwrite existing
    # If file exists, we'll overwrite it (which effectively clears it)
    if os.path.exists(output_file):
        try:
            # Check if file is writable
            if not os.access(output_file, os.W_OK):
                raise PermissionError(f"输出文件存在但不可写: '{output_file}'")
            # Remove existing file to ensure clean write
            os.remove(output_file)
            print(f"✓ 已删除现有输出文件")
        except PermissionError as e:
            print(f"错误: {str(e)}")
            sys.exit(1)
        except Exception as e:
            print(f"错误: 无法删除现有输出文件 '{output_file}': {str(e)}")
            sys.exit(1)
    
    # Write to output file with sheet name "费用"
    try:
        print(f"正在写入输出文件: '{output_file}'...")
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            final_df.to_excel(writer, sheet_name='费用', index=False, header=False)
        
        # Auto-adjust column widths after writing
        print(f"正在调整列宽...")
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, PatternFill
        
        workbook = openpyxl.load_workbook(output_file)
        worksheet = workbook['费用']
        
        # Calculate maximum width for each column
        # Check all rows including header (row 1) and data rows
        num_cols = len(final_df.columns)
        for col_idx in range(1, num_cols + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            # Check all rows (header row 1 + data rows 2 to len(final_df)+1)
            for row_idx in range(1, len(final_df) + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    cell_value = str(cell.value)
                    # For Chinese characters, multiply by 2 (rough estimate)
                    # Chinese characters are wider than ASCII characters
                    chinese_char_count = sum(1 for c in cell_value if ord(c) > 127)
                    ascii_char_count = len(cell_value) - chinese_char_count
                    # Estimate: 1 ASCII char = 1 unit, 1 Chinese char = 2 units
                    estimated_width = ascii_char_count + chinese_char_count * 2
                    if estimated_width > max_length:
                        max_length = estimated_width
            
            # Set column width with some padding (add 2 for padding)
            # Minimum width of 10, maximum width of 50
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(output_file)
        workbook.close()
        print(f"✓ 成功写入输出文件并调整列宽")
    except PermissionError:
        print(f"错误: 写入文件时权限被拒绝: '{output_file}'")
        sys.exit(1)
    except Exception as e:
        print(f"错误: 写入输出文件失败 '{output_file}': {str(e)}")
        print("请确保输出路径有效且您有写入权限。")
        sys.exit(1)
    
    # Calculate converted sums for F, G, D columns (mapped from Q, S, T)
    # Mapping: Q->F (不含税金额), S->G (税额), T->D (开票金额)
    if not result_df.empty:
        converted_q_sum = pd.to_numeric(result_df[5], errors='coerce').sum()  # F列 from Q列
        converted_s_sum = pd.to_numeric(result_df[6], errors='coerce').sum()  # G列 from S列
        converted_t_sum = pd.to_numeric(result_df[3], errors='coerce').sum()  # D列 from T列
        print(f"转换后累加值 - F列(不含税金额): {converted_q_sum:.2f}, G列(税额): {converted_s_sum:.2f}, D列(开票金额): {converted_t_sum:.2f}")
    else:
        converted_q_sum = 0
        converted_s_sum = 0
        converted_t_sum = 0
    
    # Validate sums: compare original vs converted
    tolerance = 0.01  # Allow small floating point differences
    q_match = abs(original_q_sum - converted_q_sum) < tolerance
    s_match = abs(original_s_sum - converted_s_sum) < tolerance
    t_match = abs(original_t_sum - converted_t_sum) < tolerance
    
    if q_match and s_match and t_match:
        print(f"\n✓ 验证通过: 所有累加值匹配 (Q/F, S/G, T/D)")
    else:
        print(f"\n⚠ 警告: 累加值验证失败!")
        if not q_match:
            print(f"  Q列 -> F列: 原始值={original_q_sum:.2f}, 转换后={converted_q_sum:.2f}, 差异={abs(original_q_sum - converted_q_sum):.2f}")
        if not s_match:
            print(f"  S列 -> G列: 原始值={original_s_sum:.2f}, 转换后={converted_s_sum:.2f}, 差异={abs(original_s_sum - converted_s_sum):.2f}")
        if not t_match:
            print(f"  T列 -> D列: 原始值={original_t_sum:.2f}, 转换后={converted_t_sum:.2f}, 差异={abs(original_t_sum - converted_t_sum):.2f}")
    
    print(f"\n✓ 成功处理 {len(result_df)} 行数据")
    print(f"✓ 输出已写入工作表 '费用' 到: {output_file}")
    
    # PDF Verification (if PDF directory is provided or found automatically)
    if pdf_directory:
        verify_excel_with_pdf(result_df, pdf_directory=pdf_directory, recursive=pdf_recursive)
    
    print(f"\n{'='*60}")
    print(f"✓ 转换完成!")
    print(f"  输入文件:  {input_file}")
    print(f"  输出文件: {output_file}")
    print(f"  处理行数: {len(result_df)}")
    print(f"  列累加值:")
    print(f"    Q列(不含税金额) -> F列: {original_q_sum:.2f}")
    print(f"    S列(税额) -> G列: {original_s_sum:.2f}")
    print(f"    T列(开票金额) -> D列: {original_t_sum:.2f}")
    print(f"{'='*60}")
    
    return result_df


def main():
    """Main entry point."""
    if len(sys.argv) < 2 or len(sys.argv) > 4:
        print("用法: python main.py <输入Excel文件> [输出Excel文件] [PDF目录]")
        print("示例: python main.py input.xlsx output.xlsx")
        print("示例: python main.py input.xlsx  (输出文件将在相同目录下创建为'报销.xlsx')")
        print("示例: python main.py input.xlsx output.xlsx ./pdfs  (指定PDF目录，将递归搜索所有子目录)")
        print("提示: 如果不指定PDF目录，将自动在输入文件同目录或其下的'pdfs'目录查找PDF文件（不递归）")
        print("提示: 如果指定了PDF目录，将递归搜索该目录及其所有子目录中的PDF文件")
        sys.exit(1)
    
    input_file = sys.argv[1]
    pdf_directory_specified = False  # Track if PDF directory was explicitly specified
    
    # If output file is not provided, create "报销.xlsx" in the same directory as input file
    if len(sys.argv) == 2:
        input_dir = os.path.dirname(os.path.abspath(input_file))
        if input_dir:
            output_file = os.path.join(input_dir, "报销.xlsx")
        else:
            # If input file is in current directory
            output_file = "报销.xlsx"
        print(f"未指定输出文件，使用: '{output_file}'")
        pdf_directory = None
    elif len(sys.argv) == 3:
        output_file = sys.argv[2]
        pdf_directory = None
    else:  # len(sys.argv) == 4
        output_file = sys.argv[2]
        pdf_directory = sys.argv[3]
        pdf_directory_specified = True  # PDF directory was explicitly specified
    
    # If PDF directory is not provided, try to find it automatically
    if pdf_directory is None:
        pdf_directory = find_pdf_directory(input_file)
        if pdf_directory:
            print(f"自动发现PDF目录: '{pdf_directory}'")
    
    # If PDF directory was explicitly specified, enable recursive search
    pdf_recursive = pdf_directory_specified
    
    process_excel(input_file, output_file, pdf_directory=pdf_directory, pdf_recursive=pdf_recursive)


if __name__ == "__main__":
    main()

